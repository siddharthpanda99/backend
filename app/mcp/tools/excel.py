"""MCP tools for Excel/Spreadsheet operations.

Thin wrappers over common_lib.modules.doc_processing.excel.service.
Covers workbook creation, editing, data analysis, format conversion, validation.
"""

import logging
from typing import Any, Dict, List, Optional
from mcp.server.fastmcp import FastMCP

logger = logging.getLogger("mcp.tools.excel")


def _get_svc():
    from common_lib.modules.doc_processing.excel.service import get_excel_service

    return get_excel_service()


def register_excel_tools(mcp: FastMCP):
    """Register Excel/spreadsheet tools."""

    # ── Workbook Lifecycle ──────────────────────────────────────────

    @mcp.tool()
    async def excel_create_workbook(
        sheets: Optional[List[str]] = None,
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Create a new empty Excel workbook with optional sheet names.
        Use when you need a blank workbook to populate with data."""
        try:
            svc = _get_svc()
            wb = svc.workbook.create(sheets=sheets, sheet_name=sheet_name)
            return {"workbook": str(id(wb)), "sheet_names": wb.sheetnames}
        except Exception as e:
            logger.error(f"excel_create_workbook error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_create_from_data(
        headers: List[str],
        data: List[List[Any]],
        sheet_name: str = "Data",
        auto_style: bool = False,
    ) -> Dict[str, Any]:
        """Create a workbook populated with headers and 2D data rows.
        Accepts column headers and a list of row lists."""
        try:
            wb = _get_svc().workbook.from_data(
                headers, data, sheet_name, auto_style=auto_style
            )
            return {
                "workbook": str(id(wb)),
                "sheet_names": wb.sheetnames,
                "data_rows": len(data),
            }
        except Exception as e:
            logger.error(f"excel_create_from_data error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_create_from_dicts(
        records: List[Dict[str, Any]],
        sheet_name: str = "Data",
    ) -> Dict[str, Any]:
        """Create a workbook from a list of dict records.
        Keys become headers, values become row data."""
        try:
            wb = _get_svc().workbook.from_dicts(records, sheet_name)
            return {
                "workbook": str(id(wb)),
                "sheet_names": wb.sheetnames,
                "data_rows": len(records),
            }
        except Exception as e:
            logger.error(f"excel_create_from_dicts error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_load_workbook(file_path: str) -> Dict[str, Any]:
        """Load an existing Excel workbook from a file path.
        Returns workbook ID and sheet names."""
        try:
            wb = _get_svc().io.load(file_path)
            return {
                "workbook": str(id(wb)),
                "sheet_names": wb.sheetnames,
                "path": file_path,
            }
        except Exception as e:
            logger.error(f"excel_load_workbook error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_save_workbook(file_path: str) -> Dict[str, Any]:
        """Save an existing workbook to its file path."""
        try:
            wb = _get_svc().io.load(file_path)
            _get_svc().io.save(wb, file_path)
            return {"saved": True, "path": file_path}
        except Exception as e:
            logger.error(f"excel_save_workbook error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_workbook_info(file_path: str) -> Dict[str, Any]:
        """Get metadata and structure info about a workbook."""
        try:
            return _get_svc().extractor.info(file_path)
        except Exception as e:
            logger.error(f"excel_workbook_info error: {e}")
            return {"error": str(e)}

    # ── Sheet Operations ────────────────────────────────────────────

    @mcp.tool()
    async def excel_list_sheets(file_path: str) -> Dict[str, Any]:
        """List all sheet names in a workbook."""
        try:
            wb = _get_svc().io.load(file_path)
            return {"sheet_names": wb.sheetnames, "count": len(wb.sheetnames)}
        except Exception as e:
            logger.error(f"excel_list_sheets error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_add_sheet(file_path: str, sheet_name: str) -> Dict[str, Any]:
        """Add a new sheet to a workbook."""
        try:
            svc = _get_svc()
            wb = svc.io.load(file_path)
            wb.create_sheet(sheet_name)
            svc.io.save(wb, file_path)
            return {
                "added": True,
                "sheet_name": sheet_name,
                "sheet_names": wb.sheetnames,
            }
        except Exception as e:
            logger.error(f"excel_add_sheet error: {e}")
            return {"error": str(e)}

    # ── Cell Operations ─────────────────────────────────────────────

    @mcp.tool()
    async def excel_set_cell(
        file_path: str,
        cell_ref: str,
        value: Any,
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Set a single cell value. Cell ref like 'A1', 'B2', etc."""
        try:
            svc = _get_svc()
            wb = svc.io.load(file_path)
            result = svc.cells.set(wb, cell_ref, value, sheet_name)
            svc.io.save(wb, file_path)
            return result
        except Exception as e:
            logger.error(f"excel_set_cell error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_get_cell(
        file_path: str,
        cell_ref: str,
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Get a single cell value by reference."""
        try:
            wb = _get_svc().io.load(file_path)
            return _get_svc().cells.get(wb, cell_ref, sheet_name)
        except Exception as e:
            logger.error(f"excel_get_cell error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_get_range(
        file_path: str,
        range_ref: str,
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Read a range of cells as 2D data."""
        try:
            wb = _get_svc().io.load(file_path)
            return _get_svc().cells.get_range(wb, range_ref, sheet_name)
        except Exception as e:
            logger.error(f"excel_get_range error: {e}")
            return {"error": str(e)}

    # ── Data Analysis ───────────────────────────────────────────────

    @mcp.tool()
    async def excel_analyze_summarize(data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Get summary statistics (count, mean, min, max, std) for all numeric columns in a dataset."""
        try:
            import polars as pl

            df = pl.DataFrame(data)
            result = _get_svc().statistics.describe(df)
            return {"result": result}
        except Exception as e:
            logger.error(f"excel_analyze_summarize error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_analyze_correlation(
        data: List[Dict[str, Any]],
        field_x: str,
        field_y: str,
    ) -> Dict[str, Any]:
        """Calculate Pearson correlation coefficient between two numeric fields."""
        try:
            import polars as pl

            df = pl.DataFrame(data)
            result = _get_svc().correlation.pearson(df, field_x, field_y)
            return {"result": result}
        except Exception as e:
            logger.error(f"excel_analyze_correlation error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_analyze_regression(
        data: List[Dict[str, Any]],
        field_x: str,
        field_y: str,
    ) -> Dict[str, Any]:
        """Fit a linear regression model between two numeric fields."""
        try:
            import polars as pl

            df = pl.DataFrame(data)
            result = _get_svc().regression.linear_regression(df, field_x, field_y)
            return {"result": result}
        except Exception as e:
            logger.error(f"excel_analyze_regression error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_aggregate(
        data: List[Dict[str, Any]],
        group_by: str,
        agg_field: str,
        agg_func: str = "sum",
    ) -> Dict[str, Any]:
        """Group data by a field and aggregate another field (sum, mean, count, min, max)."""
        try:
            import polars as pl

            df = pl.DataFrame(data)
            result = _get_svc().aggregate.group_by(df, group_by, agg_field, agg_func)
            return {"result": result}
        except Exception as e:
            logger.error(f"excel_aggregate error: {e}")
            return {"error": str(e)}

    # ── Validation ──────────────────────────────────────────────────

    @mcp.tool()
    async def excel_validate_data_types(
        data: List[Dict[str, Any]],
        schema: Dict[str, str],
    ) -> Dict[str, Any]:
        """Validate data types of fields against a schema dict (field_name -> expected_type).
        Types: int, float, str, number, date, email, url, phone, required."""
        try:
            result = _get_svc().data_rules.validate_data_types(data, schema)
            return result
        except Exception as e:
            logger.error(f"excel_validate_data_types error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_validate_unique(
        data: List[Dict[str, Any]],
        key_fields: List[str],
    ) -> Dict[str, Any]:
        """Check uniqueness of specified key fields across a dataset."""
        try:
            result = _get_svc().data_rules.validate_uniqueness(data, key_fields)
            return result
        except Exception as e:
            logger.error(f"excel_validate_unique error: {e}")
            return {"error": str(e)}

    # ── Format Conversion ───────────────────────────────────────────

    @mcp.tool()
    async def excel_csv_to_dicts(
        csv_string: str, delimiter: str = ","
    ) -> Dict[str, Any]:
        """Parse a CSV string into a list of dict records."""
        try:
            import csv, io

            reader = list(csv.DictReader(io.StringIO(csv_string), delimiter=delimiter))
            return {"data": reader, "count": len(reader)}
        except Exception as e:
            logger.error(f"excel_csv_to_dicts error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_dicts_to_csv(
        records: List[Dict[str, Any]],
        delimiter: str = ",",
    ) -> Dict[str, Any]:
        """Convert a list of dict records to a CSV string."""
        try:
            import csv, io

            if not records:
                return {"csv": ""}
            buf = io.StringIO()
            w = csv.DictWriter(
                buf, fieldnames=list(records[0].keys()), delimiter=delimiter
            )
            w.writeheader()
            w.writerows(records)
            return {"csv": buf.getvalue()}
        except Exception as e:
            logger.error(f"excel_dicts_to_csv error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_dicts_to_json(
        records: List[Dict[str, Any]],
        indent: int = 2,
    ) -> Dict[str, Any]:
        """Convert a list of dict records to a pretty-printed JSON string."""
        try:
            import json

            return {"json": json.dumps(records, indent=indent, default=str)}
        except Exception as e:
            logger.error(f"excel_dicts_to_json error: {e}")
            return {"error": str(e)}

    # ── Cleaning ────────────────────────────────────────────────────

    @mcp.tool()
    async def excel_deduplicate(
        data: List[Dict[str, Any]],
        key_fields: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """Remove duplicate rows from a dataset, optionally based on specific key fields."""
        try:
            import polars as pl

            df = pl.DataFrame(data)
            result = _get_svc().deduplicate.remove_duplicates(df, key_fields)
            return {"result": result}
        except Exception as e:
            logger.error(f"excel_deduplicate error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_fill_nulls(
        data: List[Dict[str, Any]],
        strategy: str = "mean",
        columns: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """Fill null/missing values. Strategies: mean, median, mode, zero, forward, backward."""
        try:
            import polars as pl

            df = pl.DataFrame(data)
            result = _get_svc().null_handling.fill_nulls(df, strategy, columns)
            return {"result": result}
        except Exception as e:
            logger.error(f"excel_fill_nulls error: {e}")
            return {"error": str(e)}

    # ── Formula Builder ─────────────────────────────────────────────

    @mcp.tool()
    async def excel_build_formula(
        formula_type: str,
        params: Dict[str, Any],
    ) -> Dict[str, Any]:
        """Build an Excel formula string. Types: sum, avg, count, min, max, if, vlookup, concat, countif, sumif, index_match."""
        try:
            svc = _get_svc()
            builders = {
                "sum": lambda: svc.formula_builder.sum(params.get("range")),
                "avg": lambda: svc.formula_builder.average(params.get("range")),
                "count": lambda: svc.formula_builder.count(params.get("range")),
                "min": lambda: svc.formula_builder.min(params.get("range")),
                "max": lambda: svc.formula_builder.max(params.get("range")),
                "if": lambda: svc.formula_builder.if_func(
                    params.get("condition"),
                    params.get("true_val"),
                    params.get("false_val"),
                ),
                "vlookup": lambda: svc.formula_builder.vlookup(
                    params.get("lookup_value"),
                    params.get("table_range"),
                    params.get("col_index"),
                    params.get("exact", False),
                ),
                "concat": lambda: svc.formula_builder.concat(params.get("cells", [])),
                "countif": lambda: svc.formula_builder.count_if(
                    params.get("range"), params.get("criteria")
                ),
                "sumif": lambda: svc.formula_builder.sum_if(
                    params.get("criteria_range"),
                    params.get("criteria"),
                    params.get("sum_range"),
                ),
                "index_match": lambda: svc.formula_builder.index_match(
                    params.get("lookup_value"),
                    params.get("index_range"),
                    params.get("match_range"),
                    params.get("exact", True),
                ),
            }
            if formula_type in builders:
                return {"formula": builders[formula_type]()}
            return {"error": f"Unknown formula type: {formula_type}"}
        except Exception as e:
            logger.error(f"excel_build_formula error: {e}")
            return {"error": str(e)}

    # ── Merge & Join ────────────────────────────────────────────────

    @mcp.tool()
    async def excel_vlookup(
        file_path: str,
        lookup_value: Any,
        table_range: str,
        col_index: int,
        exact: bool = False,
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Perform a VLOOKUP on a loaded workbook."""
        try:
            svc = _get_svc()
            wb = svc.io.load(file_path)
            result = svc.merger.vlookup(
                wb, lookup_value, table_range, col_index, exact, sheet_name
            )
            return {"result": result}
        except Exception as e:
            logger.error(f"excel_vlookup error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_pivot_data(
        data: List[Dict[str, Any]],
        index: str,
        columns: str,
        values: str,
        aggfunc: str = "sum",
    ) -> Dict[str, Any]:
        """Create a pivot table from dict data."""
        try:
            import polars as pl

            df = pl.DataFrame(data)
            result = _get_svc().pivot.pivot_data(df, index, columns, values, aggfunc)
            return {"result": result}
        except Exception as e:
            logger.error(f"excel_pivot_data error: {e}")
            return {"error": str(e)}

    # ── Render & Preview ────────────────────────────────────────────

    @mcp.tool()
    async def excel_render_sheet(
        file_path: str,
        sheet_name: str,
        max_rows: int = 100,
        max_cols: int = 20,
        include_styles: bool = True,
    ) -> Dict[str, Any]:
        """Render a worksheet as an HTML table with styling.
        Returns an HTML string suitable for embedding in dashboards or reports.
        The rendering preserves merged cells, number formats, colors, and fonts."""
        try:
            svc = _get_svc()
            wb = svc.io.load(file_path)
            result = svc.renderer.render_sheet_to_html(
                wb,
                sheet_name,
                max_rows=max_rows,
                max_cols=max_cols,
                include_styles=include_styles,
            )
            return result
        except Exception as e:
            logger.error(f"excel_render_sheet error: {e}")
            return {"error": str(e)}

    # ── Macros & VBA ────────────────────────────────────────────────

    @mcp.tool()
    async def excel_inspect_macros(file_path: str) -> Dict[str, Any]:
        """Inspect VBA macros in a workbook with risk scoring per module.
        Detects dangerous patterns (AutoOpen, Shell, CreateObject, etc.)
        and assigns a risk level (HIGH / MEDIUM / LOW) to each module."""
        try:
            svc = _get_svc()
            wb = svc.io.load(file_path)
            return svc.macro_inventory(wb)
        except Exception as e:
            logger.error(f"excel_inspect_macros error: {e}")
            return {"error": str(e)}

    # ── External References ─────────────────────────────────────────

    @mcp.tool()
    async def excel_inspect_external_links(file_path: str) -> Dict[str, Any]:
        """List external file links, references, and formula errors in a workbook.
        Returns sheet-level metadata, chart/table counts, merged cells,
        and a list of formula errors (#REF!, #DIV/0!, etc.) with locations."""
        try:
            svc = _get_svc()
            wb = svc.io.load(file_path)
            return svc.inspector.inspect_workbook(wb)
        except Exception as e:
            logger.error(f"excel_inspect_external_links error: {e}")
            return {"error": str(e)}

    # ── Calculation ─────────────────────────────────────────────────

    @mcp.tool()
    async def excel_calculate_workbook(file_path: str) -> Dict[str, Any]:
        """Force full recalculation of a workbook on next open.
        Sets calculation properties (fullCalcOnLoad) so Excel recalculates
        all formulas when the file is opened, clearing stale cached values."""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path)
            wb.calculation = openpyxl.workbook.properties.CalcProperties(
                calcMode="auto", fullCalcOnLoad=True
            )
            wb.save(file_path)
            return {
                "success": True,
                "path": file_path,
                "full_calc_on_load": True,
                "sheets": wb.sheetnames,
            }
        except Exception as e:
            logger.error(f"excel_calculate_workbook error: {e}")
            return {"error": str(e)}

    # ── Formula Audit ───────────────────────────────────────────────

    @mcp.tool()
    async def excel_audit_formulas(file_path: str) -> Dict[str, Any]:
        """Run a full formula audit across the entire workbook.
        Reports errors, circular references, inconsistent formulas,
        hardcoded constants, external references, volatile functions,
        expensive operations, and complexity scores per sheet."""
        try:
            svc = _get_svc()
            wb = svc.io.load(file_path)
            result = svc.full_audit_report(wb)
            return {"workbook": file_path, "audit": result}
        except Exception as e:
            logger.error(f"excel_audit_formulas error: {e}")
            return {"error": str(e)}

    # ── Clean Data ──────────────────────────────────────────────────

    @mcp.tool()
    async def excel_clean_data(
        file_path: str,
        sheet_name: str,
        key_fields: Optional[List[str]] = None,
        keep: str = "first",
    ) -> Dict[str, Any]:
        """Remove duplicate rows from a worksheet based on key fields.
        Loads the sheet data, deduplicates via polars, and writes the
        cleaned data back to the same sheet (overwrites existing content).
        Returns counts of original vs. cleaned rows."""
        try:
            import polars as pl
            from openpyxl.utils import get_column_letter

            svc = _get_svc()
            wb = svc.io.load(file_path)
            ws = wb[sheet_name]

            headers = [cell.value for cell in ws[1]]
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append(dict(zip(headers, row)))

            df = pl.DataFrame(data)
            original_count = len(df)
            df_clean = (
                df.unique(subset=key_fields, keep=keep)
                if key_fields
                else df.unique(keep=keep)
            )
            cleaned = df_clean.to_dicts()
            removed = original_count - len(cleaned)

            for cell in ws[1]:
                cell.value = None
            ws.delete_rows(2, ws.max_row)

            for c, h in enumerate(headers, 1):
                ws.cell(row=1, column=c, value=h)
            for r, row in enumerate(cleaned, 2):
                for c, h in enumerate(headers, 1):
                    ws.cell(row=r, column=c, value=row.get(h))

            svc.io.save(wb, file_path)
            return {
                "success": True,
                "original_rows": original_count,
                "cleaned_rows": len(cleaned),
                "duplicates_removed": removed,
                "key_fields": key_fields or "all",
            }
        except Exception as e:
            logger.error(f"excel_clean_data error: {e}")
            return {"error": str(e)}

    # ── Validate Data ───────────────────────────────────────────────

    @mcp.tool()
    async def excel_validate_data(
        file_path: str,
        data: Optional[List[Dict[str, Any]]] = None,
        schema: Optional[Dict[str, str]] = None,
        required_fields: Optional[List[str]] = None,
        unique_keys: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """Run full validation on workbook data: formula errors, structure
        issues, data type mismatches, null violations, and uniqueness checks.
        Optionally accepts external data and schema for type validation."""
        try:
            svc = _get_svc()
            wb = svc.io.load(file_path)
            result = svc.validation_runner.full_validation(
                wb,
                data=data,
                schema=schema,
                required_fields=required_fields,
                unique_keys=unique_keys,
            )
            return {"workbook": file_path, "validation": result}
        except Exception as e:
            logger.error(f"excel_validate_data error: {e}")
            return {"error": str(e)}

    # ── Transform Data ──────────────────────────────────────────────

    @mcp.tool()
    async def excel_transform_data(
        file_path: str,
        sheet_name: str,
        operation: str,
        source_col: Optional[str] = None,
        target_col: Optional[str] = None,
        separator: str = " ",
        source_cols: Optional[List[str]] = None,
        target_cols: Optional[List[str]] = None,
        id_vars: Optional[List[str]] = None,
        value_vars: Optional[List[str]] = None,
        var_name: str = "variable",
        value_name: str = "value",
        index_col: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Transform worksheet data using polars-based operations.
        Operations: split_column, merge_columns, transpose, unpivot.
        - split_column: splits source_col by separator into target_cols
        - merge_columns: joins source_cols into target_col with separator
        - transpose: swaps rows<->columns, optionally using index_col
        - unpivot: melts id_vars into value_vars with var_name/value_name
        Writes transformed data back to the sheet."""
        try:
            import polars as pl

            svc = _get_svc()
            wb = svc.io.load(file_path)
            ws = wb[sheet_name]

            headers = [cell.value for cell in ws[1]]
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append(dict(zip(headers, row)))

            df = pl.DataFrame(data)

            if operation == "split_column":
                if not source_col or not target_cols:
                    return {"error": "split_column requires source_col and target_cols"}
                result_data = svc.transform.split_column(
                    data, source_col, separator, target_cols
                )
            elif operation == "merge_columns":
                if not source_cols or not target_col:
                    return {
                        "error": "merge_columns requires source_cols and target_col"
                    }
                result_data = svc.transform.merge_columns(
                    data, source_cols, target_col, separator
                )
            elif operation == "transpose":
                result_data = svc.transform.transpose(data, index_col)
            elif operation == "unpivot":
                if not id_vars:
                    return {"error": "unpivot requires id_vars"}
                result_data = svc.transform.unpivot(
                    data, id_vars, value_vars, var_name, value_name
                )
            else:
                return {"error": f"Unknown operation: {operation}"}

            new_headers = list(result_data[0].keys()) if result_data else headers
            for cell in ws[1]:
                cell.value = None
            ws.delete_rows(2, ws.max_row)

            for c, h in enumerate(new_headers, 1):
                ws.cell(row=1, column=c, value=h)
            for r, row in enumerate(result_data, 2):
                for c, h in enumerate(new_headers, 1):
                    ws.cell(row=r, column=c, value=row.get(h))

            svc.io.save(wb, file_path)
            return {
                "success": True,
                "operation": operation,
                "rows_transformed": len(result_data),
                "columns": len(new_headers),
            }
        except Exception as e:
            logger.error(f"excel_transform_data error: {e}")
            return {"error": str(e)}

    # ── Import Data ─────────────────────────────────────────────────

    @mcp.tool()
    async def excel_import_data(
        file_path: str,
        format: str = "csv",
        sheet_name: str = "Imported",
        delimiter: str = ",",
        table_index: int = 0,
    ) -> Dict[str, Any]:
        """Import data from an external file into a new workbook.
        Supported formats: csv, json, parquet, feather, html_table, ods, xml.
        Creates a new workbook populated with the imported data and saves it."""
        try:
            import openpyxl
            import polars as pl

            svc = _get_svc()
            format_lower = format.lower()

            if format_lower == "csv":
                result = svc.from_format.from_csv_to_workbook(
                    file_path, delimiter=delimiter, sheet_name=sheet_name
                )
                wb = result["workbook"]
            elif format_lower == "json":
                raw = svc.from_format.from_json(file_path)
                df = raw["dataframe"]
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name
                for c, col in enumerate(df.columns, 1):
                    ws.cell(row=1, column=c, value=col)
                for r, row in enumerate(df.to_dicts(), 2):
                    for c, col in enumerate(df.columns, 1):
                        ws.cell(row=r, column=c, value=row[col])
            elif format_lower == "parquet":
                raw = svc.from_format.from_parquet(file_path)
                df = raw["dataframe"]
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name
                for c, col in enumerate(df.columns, 1):
                    ws.cell(row=1, column=c, value=col)
                for r, row in enumerate(df.to_dicts(), 2):
                    for c, col in enumerate(df.columns, 1):
                        ws.cell(row=r, column=c, value=row[col])
            elif format_lower == "feather":
                raw = svc.from_format.from_feather(file_path)
                df = raw["dataframe"]
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name
                for c, col in enumerate(df.columns, 1):
                    ws.cell(row=1, column=c, value=col)
                for r, row in enumerate(df.to_dicts(), 2):
                    for c, col in enumerate(df.columns, 1):
                        ws.cell(row=r, column=c, value=row[col])
            elif format_lower == "html_table":
                raw = svc.from_format.from_html_table(file_path, table_index)
                df = raw["dataframe"]
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name
                for c, col in enumerate(df.columns, 1):
                    ws.cell(row=1, column=c, value=col)
                for r, row in enumerate(df.to_dicts(), 2):
                    for c, col in enumerate(df.columns, 1):
                        ws.cell(row=r, column=c, value=row[col])
            elif format_lower == "ods":
                raw = svc.from_format.from_ods(file_path, sheet_name=None)
                if "dataframes" in raw:
                    dfs = raw["dataframes"]
                    wb = openpyxl.Workbook()
                    wb.remove(wb.active)
                    for sn, df in dfs.items():
                        ws = wb.create_sheet(title=sn[:31])
                        for c, col in enumerate(df.columns, 1):
                            ws.cell(row=1, column=c, value=col)
                        for r, row in enumerate(df.to_dicts(), 2):
                            for c, col in enumerate(df.columns, 1):
                                ws.cell(row=r, column=c, value=row[col])
                else:
                    df = raw["dataframe"]
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = sheet_name
                    for c, col in enumerate(df.columns, 1):
                        ws.cell(row=1, column=c, value=col)
                    for r, row in enumerate(df.to_dicts(), 2):
                        for c, col in enumerate(df.columns, 1):
                            ws.cell(row=r, column=c, value=row[col])
            elif format_lower == "xml":
                raw = svc.from_format.from_xml(file_path)
                df = raw["dataframe"]
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name
                for c, col in enumerate(df.columns, 1):
                    ws.cell(row=1, column=c, value=col)
                for r, row in enumerate(df.to_dicts(), 2):
                    for c, col in enumerate(df.columns, 1):
                        ws.cell(row=r, column=c, value=row[col])
            else:
                return {"error": f"Unsupported import format: {format}"}

            output_path = file_path.rsplit(".", 1)[0] + "_imported.xlsx"
            wb.save(output_path)
            return {
                "success": True,
                "output_path": output_path,
                "sheet_names": wb.sheetnames,
                "format": format_lower,
            }
        except Exception as e:
            logger.error(f"excel_import_data error: {e}")
            return {"error": str(e)}

    # ── Export Data ─────────────────────────────────────────────────

    @mcp.tool()
    async def excel_export_data(
        file_path: str,
        format: str = "csv",
        sheet_name: Optional[str] = None,
        output_path: Optional[str] = None,
        delimiter: str = ",",
    ) -> Dict[str, Any]:
        """Export a worksheet to an external format.
        Supported formats: csv, json, parquet, feather, html, markdown,
        latex, ods, pdf, xml, multi_csv, multi_json.
        Saves the exported file alongside the original by default."""
        try:
            svc = _get_svc()
            wb = svc.io.load(file_path)
            ext_map = {
                "csv": ".csv",
                "json": ".json",
                "parquet": ".parquet",
                "feather": ".feather",
                "html": ".html",
                "markdown": ".md",
                "latex": ".tex",
                "ods": ".ods",
                "pdf": ".pdf",
                "xml": ".xml",
                "multi_csv": "",
                "multi_json": ".json",
            }
            fmt = format.lower()
            if not output_path:
                base = file_path.rsplit(".", 1)[0]
                output_path = base + ext_map.get(fmt, f".{fmt}")

            if fmt == "csv":
                result = svc.to_format.to_csv(
                    wb, output_path, sheet_name, delimiter=delimiter
                )
            elif fmt == "json":
                result = svc.to_format.to_json(wb, output_path, sheet_name)
            elif fmt == "parquet":
                result = svc.to_format.to_parquet(wb, output_path, sheet_name)
            elif fmt == "feather":
                result = svc.to_format.to_feather(wb, output_path, sheet_name)
            elif fmt == "html":
                result = svc.to_format.to_html(wb, output_path, sheet_name)
            elif fmt == "markdown":
                result = svc.to_format.to_markdown(wb, output_path, sheet_name)
            elif fmt == "latex":
                result = svc.to_format.to_latex(wb, output_path, sheet_name)
            elif fmt == "ods":
                result = svc.to_format.to_ods(wb, output_path)
            elif fmt == "pdf":
                result = svc.to_format.to_pdf(wb, output_path, sheet_name)
            elif fmt == "xml":
                result = svc.to_format.to_xml(wb, output_path, sheet_name)
            elif fmt == "multi_csv":
                result = svc.to_format.to_multi_sheet_csv(
                    wb, output_path, delimiter=delimiter
                )
            elif fmt == "multi_json":
                result = svc.to_format.to_multi_sheet_json(wb, output_path)
            else:
                return {"error": f"Unsupported export format: {fmt}"}
            return {"success": True, **result}
        except Exception as e:
            logger.error(f"excel_export_data error: {e}")
            return {"error": str(e)}

    # ── Convert File ────────────────────────────────────────────────

    @mcp.tool()
    async def excel_convert_file(
        file_path: str,
        target_format: str,
        output_path: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Convert a tabular file from one format to another.
        Auto-detects source format from the file extension.
        Supports: csv, json, parquet, feather.
        Does not require a workbook intermediary."""
        try:
            svc = _get_svc()
            if not output_path:
                base = file_path.rsplit(".", 1)[0]
                output_path = f"{base}.{target_format}"
            result = svc.converters.auto_detect_and_convert(file_path, output_path)
            return {
                "success": True,
                "input": file_path,
                "output": output_path,
                **result,
            }
        except Exception as e:
            logger.error(f"excel_convert_file error: {e}")
            return {"error": str(e)}

    # ── Extract Tables ──────────────────────────────────────────────

    @mcp.tool()
    async def excel_extract_tables(
        file_path: str, sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """Detect and extract tabular regions from a worksheet using
        AI-driven heuristics. Identifies contiguous data blocks with
        headers, row/column counts, and cell ranges.
        When sheet_name is omitted, scans ALL sheets."""
        try:
            svc = _get_svc()
            wb = svc.io.load(file_path)
            return svc.detect_tables(wb, sheet_name)
        except Exception as e:
            logger.error(f"excel_extract_tables error: {e}")
            return {"error": str(e)}

    # ── Extract Comments ────────────────────────────────────────────

    @mcp.tool()
    async def excel_extract_comments(
        file_path: str, sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """Extract all comments/notes from a worksheet.
        Returns cell address, sheet name, comment text, author,
        and resolved status for each comment."""
        try:
            svc = _get_svc()
            wb = svc.io.load(file_path)
            comments = svc.comments.list_comments(wb, sheet_name)
            return {
                "success": True,
                "comment_count": len(comments),
                "comments": comments,
                "sheet": sheet_name or "all",
            }
        except Exception as e:
            logger.error(f"excel_extract_comments error: {e}")
            return {"error": str(e)}

    # ── Move Range ──────────────────────────────────────────────────

    @mcp.tool()
    async def excel_move_range(
        file_path: str,
        range_ref: str,
        destination: str,
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Move a cell range to a new location (copy + clear original).
        Copies all values and formats from range_ref to the destination
        cell, then clears the original range contents."""
        try:
            svc = _get_svc()
            wb = svc.io.load(file_path)
            ws = wb[sheet_name] if sheet_name else wb.active
            from openpyxl.utils import range_boundaries

            min_col, min_row, max_col, max_row = range_boundaries(range_ref)
            data = []
            for r in range(min_row, max_row + 1):
                row_data = []
                for c in range(min_col, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    row_data.append(
                        {
                            "value": cell.value,
                            "number_format": cell.number_format,
                            "font": cell.font,
                            "fill": cell.fill,
                            "border": cell.border,
                            "alignment": cell.alignment,
                        }
                    )
                data.append(row_data)

            from openpyxl.utils import column_index_from_string, get_column_letter

            dest_cell = ws[destination]
            dr, dc = dest_cell.row, dest_cell.column
            rows = len(data)
            cols = len(data[0]) if data else 0
            for r in range(rows):
                for c in range(cols):
                    target = ws.cell(row=dr + r, column=dc + c)
                    src = data[r][c]
                    target.value = src["value"]
                    target.number_format = src["number_format"]
                    target.font = src["font"]
                    target.fill = src["fill"]
                    target.border = src["border"]
                    target.alignment = src["alignment"]

            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.value = None

            svc.io.save(wb, file_path)
            return {
                "success": True,
                "source": range_ref,
                "destination": destination,
                "rows": rows,
                "cols": cols,
            }
        except Exception as e:
            logger.error(f"excel_move_range error: {e}")
            return {"error": str(e)}

    # ── Sort & Filter ────────────────────────────────────────────────

    @mcp.tool()
    async def excel_sort_range(
        file_path: str,
        sheet_name: str,
        range_address: str,
        sort_by_column: int = 1,
        descending: bool = False,
    ) -> Dict[str, Any]:
        """Sort a cell range in a worksheet by a specific column.
        Preserves the header row (row 1 of the range)."""
        try:
            svc = _get_svc()
            wb = svc.io.load(file_path)
            result = svc.sort.sort_range(
                wb,
                range_address,
                key_column=sort_by_column,
                ascending=not descending,
                sheet_name=sheet_name,
            )
            svc.io.save(wb, file_path)
            return {"success": result["sorted"], "rows_affected": result["rows"]}
        except Exception as e:
            logger.error(f"excel_sort_range error: {e}")
            return {"error": str(e)}

    @mcp.tool()
    async def excel_filter_range(
        file_path: str,
        sheet_name: str,
        range_address: str,
        field_index: int,
        criteria: List[Any],
    ) -> Dict[str, Any]:
        """Filter rows in a worksheet range where the specified column matches ANY of the given criteria values.
        Skips the header row. Returns the count of matching rows."""
        try:
            from openpyxl.worksheet.cell_range import CellRange

            wb = _get_svc().io.load(file_path)
            ws = wb[sheet_name]
            cr = CellRange(range_address)
            criteria_set = set(criteria)
            col_idx = field_index - 1
            count = 0
            for row in ws.iter_rows(
                min_row=cr.min_row + 1,
                max_row=cr.max_row,
                min_col=cr.min_col,
                max_col=cr.max_col,
                values_only=True,
            ):
                if len(row) > col_idx and row[col_idx] in criteria_set:
                    count += 1
            return {"success": True, "filtered_rows": count}
        except Exception as e:
            logger.error(f"excel_filter_range error: {e}")
            return {"error": str(e)}

    # ── Versioning ───────────────────────────────────────────────────

    @mcp.tool()
    async def excel_compare_versions(
        file_path_v1: str,
        file_path_v2: str,
        sheet_name: str,
    ) -> Dict[str, Any]:
        """Compare two workbook files cell-by-cell on a specific sheet.
        Returns a list of differences with cell addresses, old values, and new values."""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter

            wb1 = openpyxl.load_workbook(file_path_v1, data_only=True)
            wb2 = openpyxl.load_workbook(file_path_v2, data_only=True)
            ws1 = wb1[sheet_name] if sheet_name in wb1.sheetnames else None
            ws2 = wb2[sheet_name] if sheet_name in wb2.sheetnames else None
            if not ws1 and not ws2:
                return {"error": f"Sheet '{sheet_name}' not found in either workbook"}
            max_row = max(ws1.max_row if ws1 else 0, ws2.max_row if ws2 else 0)
            max_col = max(ws1.max_column if ws1 else 0, ws2.max_column if ws2 else 0)
            differences = []
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    val1 = ws1.cell(row=r, column=c).value if ws1 else None
                    val2 = ws2.cell(row=r, column=c).value if ws2 else None
                    if val1 != val2:
                        differences.append(
                            {
                                "cell_address": f"{get_column_letter(c)}{r}",
                                "old_value": val1,
                                "new_value": val2,
                            }
                        )
            return {"differences": differences, "change_count": len(differences)}
        except Exception as e:
            logger.error(f"excel_compare_versions error: {e}")
            return {"error": str(e)}

    logger.info("Excel: %s MCP tools registered", 35)
