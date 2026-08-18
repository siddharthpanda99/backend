"""Excel/Spreadsheet REST API routes — thin routing layer over common_lib ExcelService.

All business logic lives in common_lib.modules.doc_processing.excel.service.
This file is the HTTP facade: parse request, call service, return response.
"""

from __future__ import annotations

import logging
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, HTTPException, Query
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

router = APIRouter()


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------


class FilePathRequest(BaseModel):
    file_path: str


class CreateWorkbookRequest(BaseModel):
    sheets: Optional[List[str]] = None
    sheet_name: Optional[str] = None


class FromDataRequest(BaseModel):
    headers: List[str]
    data: List[List[Any]]
    sheet_name: Optional[str] = "Data"
    title: Optional[str] = None
    auto_style: Optional[bool] = False


class FromDictsRequest(BaseModel):
    records: List[Dict[str, Any]]
    sheet_name: Optional[str] = "Data"


class SheetRequest(BaseModel):
    file_path: str
    sheet_name: str


class RenameSheetRequest(BaseModel):
    file_path: str
    old_name: str
    new_name: str


class CellRequest(BaseModel):
    file_path: str
    cell_ref: str
    sheet_name: Optional[str] = None


class SetCellRequest(BaseModel):
    file_path: str
    cell_ref: str
    value: Any
    sheet_name: Optional[str] = None


class RangeRequest(BaseModel):
    file_path: str
    range_ref: str
    sheet_name: Optional[str] = None


class SetRangeRequest(BaseModel):
    file_path: str
    range_ref: str
    data: List[List[Any]]
    sheet_name: Optional[str] = None


class BatchCellRequest(BaseModel):
    file_path: str
    data: List[Dict[str, Any]]
    sheet_name: Optional[str] = None


class StyleRequest(BaseModel):
    file_path: str
    range_ref: str
    style: Dict[str, Any]
    sheet_name: Optional[str] = None


class MergeRequest(BaseModel):
    file_path: str
    range_ref: str
    sheet_name: Optional[str] = None


class FreezeRequest(BaseModel):
    file_path: str
    cell_ref: str
    sheet_name: Optional[str] = None


class RowColRequest(BaseModel):
    file_path: str
    row: Optional[int] = None
    column: Optional[int] = None
    sheet_name: Optional[str] = None


class ConditionalFormatRequest(BaseModel):
    file_path: str
    range_ref: str
    rule: Dict[str, Any]
    sheet_name: Optional[str] = None


class TableRequest(BaseModel):
    file_path: str
    range_ref: str
    table_name: Optional[str] = None
    sheet_name: Optional[str] = None


class ChartRequest(BaseModel):
    file_path: str
    chart_type: str
    range_ref: str
    title: Optional[str] = None
    sheet_name: Optional[str] = None


class AnalysisRequest(BaseModel):
    data: List[Dict[str, Any]]
    group_by: Optional[str] = None
    agg_field: Optional[str] = None
    agg_func: Optional[str] = "sum"
    agg_fields: Optional[List[Dict[str, str]]] = None
    field: Optional[str] = None
    field_x: Optional[str] = None
    field_y: Optional[str] = None
    bins: Optional[int] = 10
    window: Optional[int] = 3
    target: Optional[float] = None
    variable_range: Optional[Dict[str, Any]] = None
    scenarios: Optional[List[Dict[str, Any]]] = None
    func: Optional[str] = None
    filters: Optional[Dict[str, Any]] = None
    min_val: Optional[float] = None
    max_val: Optional[float] = None
    n: Optional[int] = 10
    ascending: Optional[bool] = False
    sort_fields: Optional[List[Dict[str, Any]]] = None
    query: Optional[str] = None
    mode: Optional[str] = "contains"
    model: Optional[Dict[str, Any]] = None
    x_values: Optional[List[float]] = None
    base_data: Optional[List[Dict[str, Any]]] = None


class CleanRequest(BaseModel):
    data: List[Dict[str, Any]]
    key_fields: Optional[List[str]] = None
    field: Optional[str] = None
    strategy: Optional[str] = "mean"
    method: Optional[str] = "minmax"
    columns: Optional[List[str]] = None
    threshold: Optional[float] = None


class ValidateRequest(BaseModel):
    file_path: Optional[str] = None
    data: Optional[List[Dict[str, Any]]] = None
    type_schema: Optional[Dict[str, str]] = Field(None, alias="schema")
    required_fields: Optional[List[str]] = None
    unique_keys: Optional[List[str]] = None
    key_fields: Optional[List[str]] = None
    field: Optional[str] = None
    min_val: Optional[float] = None
    max_val: Optional[float] = None
    allowed_values: Optional[List[Any]] = None
    expected_sheets: Optional[List[str]] = None
    sheet_name: Optional[str] = None


class MergeWorkbooksRequest(BaseModel):
    file_paths: List[str]
    output_path: Optional[str] = None


class VLookupRequest(BaseModel):
    file_path: str
    lookup_value: Any
    table_range: str
    col_index: int
    exact: Optional[bool] = False
    sheet_name: Optional[str] = None


class PivotRequest(BaseModel):
    data: List[Dict[str, Any]]
    index: str
    columns: str
    values: str
    aggfunc: Optional[str] = "sum"


class ConvertRequest(BaseModel):
    input_path: str
    output_path: str
    sheet_name: Optional[str] = None


class WorkbookConvertRequest(BaseModel):
    workbook_data: Optional[str] = None  # base64
    file_path: Optional[str] = None
    sheet_name: Optional[str] = None


class DataConvertRequest(BaseModel):
    records: List[Dict[str, Any]]
    indent: Optional[int] = 2
    delimiter: Optional[str] = ","
    headers: Optional[List[str]] = None


class StringConvertRequest(BaseModel):
    string: str
    delimiter: Optional[str] = ","


class BuildFormulaRequest(BaseModel):
    formula_type: str  # sum, avg, count, min, max, if, vlookup, concat, countif, sumif, index_match
    params: Dict[str, Any]


class NamedRangeRequest(BaseModel):
    file_path: str
    range_name: str
    range_ref: Optional[str] = None
    sheet_name: Optional[str] = None


class CommentRequest(BaseModel):
    file_path: str
    cell_ref: str
    text: str
    author: Optional[str] = None
    sheet_name: Optional[str] = None


class HyperlinkRequest(BaseModel):
    file_path: str
    cell_ref: str
    url: str
    text: Optional[str] = None
    sheet_name: Optional[str] = None


class ImageRequest(BaseModel):
    file_path: str
    image_path: str
    cell_ref: str
    width: Optional[int] = None
    height: Optional[int] = None
    sheet_name: Optional[str] = None


class PrintLayoutRequest(BaseModel):
    file_path: str
    sheet_name: Optional[str] = None
    settings: Dict[str, Any]


class SecurityRequest(BaseModel):
    file_path: str
    password: Optional[str] = None
    sheet_name: Optional[str] = None
    protection_settings: Optional[Dict[str, bool]] = None


class VersionRequest(BaseModel):
    file_path: str
    version_path: Optional[str] = None
    comment: Optional[str] = None


class AIOpsRequest(BaseModel):
    file_path: str
    operation: str
    params: Dict[str, Any]


class SortRangeRequest(BaseModel):
    file_path: str
    sheet_name: str
    range_address: str
    sort_by_column: int = Field(default=1, ge=1)
    descending: bool = False


class FilterRangeRequest(BaseModel):
    file_path: str
    sheet_name: str
    range_address: str
    field_index: int = Field(..., ge=1)
    criteria: List[Any] = []


class SnapshotRequest(BaseModel):
    file_path: str
    label: str


class DiffRequest(BaseModel):
    file_path: str
    snapshot_id_1: str
    snapshot_id_2: str


class AsyncConvertRequest(BaseModel):
    target_format: str
    options: Optional[Dict[str, Any]] = None


class AsyncOcrRequest(BaseModel):
    image_path: str
    options: Optional[Dict[str, Any]] = None


class AsyncExportRequest(BaseModel):
    export_format: str
    options: Optional[Dict[str, Any]] = None


class RenderSheetRequest(BaseModel):
    file_path: str
    sheet_name: str = "Sheet1"
    max_rows: Optional[int] = 100
    max_cols: Optional[int] = 20
    include_styles: Optional[bool] = True


class RenderMarkdownRequest(BaseModel):
    file_path: str
    sheet_name: Optional[str] = None
    max_rows: Optional[int] = None


class SecurityScanRequest(BaseModel):
    file_path: str
    deep: Optional[bool] = False


class OcrExtractRequest(BaseModel):
    file_path: str
    sheet_name: Optional[str] = "Sheet1"


class CollabLockRequest(BaseModel):
    workbook_id: str
    cell_ref: str
    user_id: str
    sheet_name: str


class CollabUnlockRequest(BaseModel):
    lock_id: str
    user_id: str


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _get_svc():
    from common_lib.modules.doc_processing.excel.service import get_excel_service

    return get_excel_service()


def _load_wb(file_path: str):
    return _get_svc().io.load(file_path)


def _save_wb(workbook, file_path: str):
    _get_svc().io.save(workbook, file_path)
    return {"saved": True, "path": file_path}


# ---------------------------------------------------------------------------
# Workbook Lifecycle
# ---------------------------------------------------------------------------


@router.post("/excel/create")
async def create_workbook(
    req: CreateWorkbookRequest,
    dry_run: bool = Query(False),
) -> Dict[str, Any]:
    if dry_run:
        sheets = req.sheets or [req.sheet_name or "Sheet"]
        return {
            "dry_run": True,
            "estimated_impact": {
                "action": "create_workbook",
                "sheets": sheets,
                "sheet_count": len(sheets),
            },
        }
    try:
        svc = _get_svc()
        wb = svc.workbook.create(sheets=req.sheets, sheet_name=req.sheet_name)
        return {"workbook": str(id(wb)), "sheet_names": wb.sheetnames}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/from-data")
async def create_from_data(req: FromDataRequest) -> Dict[str, Any]:
    try:
        wb = _get_svc().workbook.from_data(
            req.headers, req.data, req.sheet_name, req.title, req.auto_style
        )
        return {
            "workbook": str(id(wb)),
            "sheet_names": wb.sheetnames,
            "data_rows": len(req.data),
        }
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/from-dicts")
async def create_from_dicts(req: FromDictsRequest) -> Dict[str, Any]:
    try:
        wb = _get_svc().workbook.from_dicts(req.records, req.sheet_name)
        return {
            "workbook": str(id(wb)),
            "sheet_names": wb.sheetnames,
            "data_rows": len(req.records),
        }
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/load")
async def load_workbook(req: FilePathRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        return {
            "workbook": str(id(wb)),
            "sheet_names": wb.sheetnames,
            "path": req.file_path,
        }
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/save")
async def save_workbook(req: FilePathRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        _save_wb(wb, req.file_path)
        return {"saved": True, "path": req.file_path}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/info")
async def workbook_info(req: FilePathRequest) -> Dict[str, Any]:
    try:
        return _get_svc().extractor.info(req.file_path)
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Sheet Operations
# ---------------------------------------------------------------------------


@router.post("/excel/sheet/list")
async def list_sheets(req: FilePathRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        return {"sheet_names": wb.sheetnames, "count": len(wb.sheetnames)}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/sheet/add")
async def add_sheet(req: SheetRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        wb.create_sheet(req.sheet_name)
        _save_wb(wb, req.file_path)
        return {
            "added": True,
            "sheet_name": req.sheet_name,
            "sheet_names": wb.sheetnames,
        }
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/sheet/rename")
async def rename_sheet(req: RenameSheetRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        ws = wb[req.old_name]
        ws.title = req.new_name
        _save_wb(wb, req.file_path)
        return {"renamed": True, "from": req.old_name, "to": req.new_name}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/sheet/delete")
async def delete_sheet(req: SheetRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        del wb[req.sheet_name]
        _save_wb(wb, req.file_path)
        return {"deleted": True, "sheet_name": req.sheet_name}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/sheet/hide")
async def hide_sheet(req: SheetRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.rows_cols.hide_sheet(wb, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/sheet/unhide")
async def unhide_sheet(req: SheetRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.rows_cols.unhide_sheet(wb, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Cell Operations
# ---------------------------------------------------------------------------


@router.post("/excel/cell/set")
async def set_cell(req: SetCellRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.cells.set(wb, req.cell_ref, req.value, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/cell/get")
async def get_cell(req: CellRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        result = _get_svc().cells.get(wb, req.cell_ref, req.sheet_name)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/cells/batch")
async def batch_set_cells(
    req: BatchCellRequest,
    dry_run: bool = Query(False),
) -> Dict[str, Any]:
    if dry_run:
        return {
            "dry_run": True,
            "estimated_impact": {
                "action": "batch_set_cells",
                "cell_count": len(req.data),
                "sheet": req.sheet_name or "active",
            },
        }
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.cells.batch_set(wb, req.data, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Range Operations
# ---------------------------------------------------------------------------


@router.post("/excel/range/set")
async def set_range(req: SetRangeRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.cells.set_range(wb, req.range_ref, req.data, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/range/get")
async def get_range(req: RangeRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        result = _get_svc().cells.get_range(wb, req.range_ref, req.sheet_name)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/range/copy")
async def copy_range(req: RangeRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.cells.copy_range(wb, req.range_ref, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/range/read-data")
async def read_range_data(
    req: RangeRequest,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=1000),
) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        result = _get_svc().extractor.read_range(wb, req.range_ref, req.sheet_name)
        rows = result.get("rows", [])
        total = len(rows)
        total_pages = max(1, (total + page_size - 1) // page_size)
        start = (page - 1) * page_size
        end = start + page_size
        return {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": total_pages,
            "data": rows[start:end],
            **{k: v for k, v in result.items() if k != "rows"},
        }
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Row / Column Operations
# ---------------------------------------------------------------------------


@router.post("/excel/row/insert")
async def insert_row(req: RowColRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.rows_cols.insert_row(wb, req.row, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/row/delete")
async def delete_row(req: RowColRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.rows_cols.delete_row(wb, req.row, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/column/insert")
async def insert_column(req: RowColRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.rows_cols.insert_column(wb, req.column, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/column/delete")
async def delete_column(req: RowColRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.rows_cols.delete_column(wb, req.column, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Formatting
# ---------------------------------------------------------------------------


@router.post("/excel/format/style")
async def set_style(req: StyleRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.formatting.apply_style(
            wb, req.range_ref, req.style, req.sheet_name
        )
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/format/merge")
async def merge_cells(req: MergeRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.merge_freeze.merge(wb, req.range_ref, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/format/unmerge")
async def unmerge_cells(req: MergeRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.merge_freeze.unmerge(wb, req.range_ref, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/format/freeze")
async def freeze_panes(req: FreezeRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.merge_freeze.freeze(wb, req.cell_ref, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/format/auto-fit")
async def auto_fit(req: FilePathRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.formatting.auto_fit(wb)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/format/column-width")
async def set_column_width(req: SheetRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.rows_cols.set_column_width(wb, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/format/conditional")
async def conditional_format(req: ConditionalFormatRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.formatting.apply_conditional(
            wb, req.range_ref, req.rule, req.sheet_name
        )
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Tables & Charts
# ---------------------------------------------------------------------------


@router.post("/excel/table/add")
async def add_table(req: TableRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        result = svc.tables.add(
            req.file_path, req.range_ref, req.table_name, req.sheet_name
        )
        wb = _load_wb(req.file_path)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/chart/add")
async def add_chart(req: ChartRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        result = svc.charts.add(
            req.file_path, req.chart_type, req.range_ref, req.title, req.sheet_name
        )
        wb = _load_wb(req.file_path)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Data Export
# ---------------------------------------------------------------------------


@router.post("/excel/export/json")
async def export_to_json(req: WorkbookConvertRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        svc = _get_svc()
        data = svc.extractor.to_dicts(wb, req.sheet_name)
        import json

        return {"data": json.dumps(data, indent=2, default=str)}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/export/csv")
async def export_to_csv(req: WorkbookConvertRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        svc = _get_svc()
        result = svc.extractor.to_csv(wb, req.sheet_name)
        return {"csv": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/export/markdown")
async def export_to_markdown(req: WorkbookConvertRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        svc = _get_svc()
        result = svc.extractor.to_markdown(wb, req.sheet_name)
        return {"markdown": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/export/html")
async def export_to_html(req: WorkbookConvertRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        svc = _get_svc()
        result = svc.extractor.to_html(wb, req.sheet_name)
        return {"html": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Data Import
# ---------------------------------------------------------------------------


@router.post("/excel/import/csv")
async def import_csv(req: WorkbookConvertRequest) -> Dict[str, Any]:
    try:
        if req.file_path and req.workbook_data:
            svc = _get_svc()
            wb = svc.csv_engine.to_workbook(req.workbook_data, req.sheet_name)
            return {"workbook": str(id(wb)), "sheet_names": wb.sheetnames}
        raise HTTPException(
            400,
            detail="file_path not supported for CSV import, pass csv_string as workbook_data",
        )
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/import/json")
async def import_json(req: WorkbookConvertRequest) -> Dict[str, Any]:
    try:
        if req.workbook_data:
            import json

            records = json.loads(req.workbook_data)
            wb = _get_svc().workbook.from_dicts(records, req.sheet_name or "Imported")
            return {"workbook": str(id(wb)), "sheet_names": wb.sheetnames}
        raise HTTPException(400, detail="json_string required as workbook_data")
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Data Conversion
# ---------------------------------------------------------------------------


@router.post("/excel/convert/csv-to-json")
async def csv_to_json(
    req: StringConvertRequest,
    dry_run: bool = Query(False),
) -> Dict[str, Any]:
    if dry_run:
        import csv, io

        preview = list(
            csv.DictReader(io.StringIO(req.string[:2048]), delimiter=req.delimiter)
        )
        total_estimate = max(0, req.string.count("\n"))
        return {
            "dry_run": True,
            "estimated_impact": {
                "action": "csv_to_json",
                "input_length": len(req.string),
                "estimated_records": total_estimate,
                "preview_fields": list(preview[0].keys()) if preview else [],
            },
        }
    try:
        import json
        import csv, io

        reader = list(csv.DictReader(io.StringIO(req.string), delimiter=req.delimiter))
        return {"json": json.dumps(reader, indent=2, default=str)}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/convert/json-to-csv")
async def json_to_csv(req: StringConvertRequest) -> Dict[str, Any]:
    try:
        import json, csv, io

        records = json.loads(req.string)
        if not records:
            return {"csv": ""}
        records = records if isinstance(records, list) else [records]
        buf = io.StringIO()
        w = csv.DictWriter(
            buf, fieldnames=list(records[0].keys()), delimiter=req.delimiter
        )
        w.writeheader()
        w.writerows(records)
        return {"csv": buf.getvalue()}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/convert/csv-to-dicts")
async def csv_to_dicts(req: StringConvertRequest) -> Dict[str, Any]:
    try:
        import csv, io

        reader = list(csv.DictReader(io.StringIO(req.string), delimiter=req.delimiter))
        return {"data": reader, "count": len(reader)}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/convert/dicts-to-csv")
async def dicts_to_csv(req: DataConvertRequest) -> Dict[str, Any]:
    try:
        import csv, io

        if not req.records:
            return {"csv": ""}
        buf = io.StringIO()
        w = csv.DictWriter(
            buf, fieldnames=list(req.records[0].keys()), delimiter=req.delimiter
        )
        w.writeheader()
        w.writerows(req.records)
        return {"csv": buf.getvalue()}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/convert/dicts-to-json")
async def dicts_to_json(req: DataConvertRequest) -> Dict[str, Any]:
    try:
        import json

        return {"json": json.dumps(req.records, indent=req.indent, default=str)}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/convert/json-to-dicts")
async def json_to_dicts(req: StringConvertRequest) -> Dict[str, Any]:
    try:
        import json

        records = json.loads(req.string)
        if isinstance(records, dict):
            records = [records]
        return {"data": records, "count": len(records)}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/convert/dicts-to-html")
async def dicts_to_html(req: DataConvertRequest) -> Dict[str, Any]:
    try:
        if not req.records:
            return {"html": ""}
        h = list(req.records[0].keys())
        thead = "".join(f"<th>{k}</th>" for k in h)
        rows = []
        for r in req.records:
            cells = "".join(f"<td>{r.get(k, '')}</td>" for k in h)
            rows.append(f"<tr>{cells}</tr>")
        tbody = "\n".join(rows)
        return {
            "html": f'<table border="1" cellpadding="4" cellspacing="0"><thead><tr>{thead}</tr></thead><tbody>{tbody}</tbody></table>'
        }
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/convert/dicts-to-markdown")
async def dicts_to_markdown(req: DataConvertRequest) -> Dict[str, Any]:
    try:
        if not req.records:
            return {"markdown": ""}
        h = list(req.records[0].keys())
        lines = [
            "| " + " | ".join(h) + " |",
            "| " + " | ".join(["---"] * len(h)) + " |",
        ]
        for r in req.records:
            lines.append("| " + " | ".join(str(r.get(k, "")) for k in h) + " |")
        return {"markdown": "\n".join(lines)}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Data Analysis
# ---------------------------------------------------------------------------


@router.post("/excel/analyze/aggregate")
async def aggregate(req: AnalysisRequest) -> Dict[str, Any]:
    try:
        import polars as pl

        df = pl.DataFrame(req.data)
        svc = _get_svc()
        if req.agg_fields:
            result = svc.aggregate.multi_aggregate(df, req.group_by, req.agg_fields)
        else:
            result = svc.aggregate.group_by(
                df, req.group_by, req.agg_field, req.agg_func
            )
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/analyze/summarize")
async def summarize(req: AnalysisRequest) -> Dict[str, Any]:
    try:
        import polars as pl

        df = pl.DataFrame(req.data)
        result = _get_svc().statistics.describe(df)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/analyze/describe")
async def describe(req: AnalysisRequest) -> Dict[str, Any]:
    try:
        import polars as pl

        df = pl.DataFrame(req.data)
        result = _get_svc().statistics.describe_all(df)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/analyze/frequency")
async def frequency(req: AnalysisRequest) -> Dict[str, Any]:
    try:
        import polars as pl

        df = pl.DataFrame(req.data)
        result = _get_svc().statistics.frequency(df, req.field)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/analyze/histogram")
async def histogram(req: AnalysisRequest) -> Dict[str, Any]:
    try:
        import polars as pl

        df = pl.DataFrame(req.data)
        result = _get_svc().statistics.histogram(df, req.field, req.bins)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/analyze/correlation")
async def correlation(req: AnalysisRequest) -> Dict[str, Any]:
    try:
        import polars as pl

        df = pl.DataFrame(req.data)
        result = _get_svc().correlation.pearson(df, req.field_x, req.field_y)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/analyze/regression")
async def linear_regression(req: AnalysisRequest) -> Dict[str, Any]:
    try:
        import polars as pl

        df = pl.DataFrame(req.data)
        result = _get_svc().regression.linear_regression(df, req.field_x, req.field_y)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/analyze/forecast")
async def moving_average(req: AnalysisRequest) -> Dict[str, Any]:
    try:
        import polars as pl

        df = pl.DataFrame(req.data)
        result = _get_svc().forecasting.simple_moving_average(df, req.field, req.window)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/analyze/what-if")
async def what_if(req: AnalysisRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        if req.func:
            import polars as pl

            df = pl.DataFrame(req.data)
            result = svc.what_if.goal_seek(
                eval(req.func), req.target, req.variable_range
            )
            return {"result": result}
        elif req.scenarios and req.base_data:
            result = svc.what_if.scenario_manager(req.base_data, req.scenarios)
            return {"result": result}
        raise HTTPException(
            400,
            detail="Provide either func+target+variable_range or base_data+scenarios",
        )
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/analyze/stats")
async def compute_stats(req: AnalysisRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        result = _get_svc().statistics.compute_stats(req.data, req.field)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/analyze/filter")
async def filter_data(req: AnalysisRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        result = svc.statistics.filter_by(req.data, req.filters)
        return {"data": result, "count": len(result)}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/analyze/top-n")
async def top_n(req: AnalysisRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        result = svc.statistics.top_n(req.data, req.field, req.n, req.ascending)
        return {"data": result, "count": len(result)}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/analyze/unique")
async def unique_values(req: AnalysisRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        result = svc.statistics.unique_values(req.data, req.field)
        return {"values": result, "count": len(result)}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/analyze/sort")
async def sort_data(req: AnalysisRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        if req.sort_fields:
            result = svc.statistics.sort_by_multiple(req.data, req.sort_fields)
        else:
            result = svc.statistics.sort_by(req.data, req.field, req.ascending)
        return {"data": result, "count": len(result)}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Data Cleaning
# ---------------------------------------------------------------------------


@router.post("/excel/clean/deduplicate")
async def deduplicate(req: CleanRequest) -> Dict[str, Any]:
    try:
        import polars as pl

        df = pl.DataFrame(req.data)
        result = _get_svc().deduplicate.remove_duplicates(df, req.key_fields)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/clean/normalize")
async def normalize(req: CleanRequest) -> Dict[str, Any]:
    try:
        import polars as pl

        df = pl.DataFrame(req.data)
        result = _get_svc().normalize.normalize_columns(df, req.method, req.columns)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/clean/fill-nulls")
async def fill_nulls(req: CleanRequest) -> Dict[str, Any]:
    try:
        import polars as pl

        df = pl.DataFrame(req.data)
        result = _get_svc().null_handling.fill_nulls(df, req.strategy, req.columns)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/clean/outliers")
async def detect_outliers(req: CleanRequest) -> Dict[str, Any]:
    try:
        import polars as pl

        df = pl.DataFrame(req.data)
        result = _get_svc().outlier.detect_outliers(df, req.field, req.threshold)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Data Validation
# ---------------------------------------------------------------------------


@router.post("/excel/validate/formulas")
async def validate_formulas(req: ValidateRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        result = _get_svc().formula_audit.scan_errors(wb)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/validate/data-types")
async def validate_data_types(req: ValidateRequest) -> Dict[str, Any]:
    try:
        result = _get_svc().data_rules.validate_data_types(req.data, req.type_schema)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/validate/unique")
async def validate_unique(req: ValidateRequest) -> Dict[str, Any]:
    try:
        result = _get_svc().data_rules.validate_uniqueness(req.data, req.key_fields)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/validate/not-null")
async def validate_not_null(req: ValidateRequest) -> Dict[str, Any]:
    try:
        result = _get_svc().data_rules.validate_not_null(req.data, req.required_fields)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/validate/range")
async def validate_range(req: ValidateRequest) -> Dict[str, Any]:
    try:
        result = _get_svc().data_rules.validate_range(
            req.data, req.field, req.min_val, req.max_val
        )
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/validate/enum")
async def validate_enum(req: ValidateRequest) -> Dict[str, Any]:
    try:
        result = _get_svc().data_rules.validate_enum(
            req.data, req.field, req.allowed_values
        )
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/validate/structure")
async def validate_structure(req: ValidateRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        result = _get_svc().structure.validate_workbook_structure(wb)
        if req.expected_sheets:
            missing = set(req.expected_sheets) - set(wb.sheetnames)
            if missing:
                result.setdefault("issues", []).append(
                    {"issue": "missing_sheets", "sheets": list(missing)}
                )
        result["valid"] = len(result.get("issues", [])) == 0
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/validate/full")
async def full_validation(req: ValidateRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        result = _get_svc().full_check.run(
            wb, req.data, req.type_schema, req.required_fields, req.unique_keys
        )
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/validate/circular")
async def check_circular(req: ValidateRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        result = _get_svc().formula_audit.find_circular_references(wb)
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Merge & Join Operations
# ---------------------------------------------------------------------------


@router.post("/excel/merge/workbooks")
async def merge_workbooks(req: MergeWorkbooksRequest) -> Dict[str, Any]:
    try:
        result = _get_svc().merger.merge_workbooks(req.file_paths, req.output_path)
        return {"result": result, "output": req.output_path}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/merge/vlookup")
async def vlookup(req: VLookupRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        result = _get_svc().merger.vlookup(
            wb,
            req.lookup_value,
            req.table_range,
            req.col_index,
            req.exact,
            req.sheet_name,
        )
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/merge/pivot")
async def pivot_data(req: PivotRequest) -> Dict[str, Any]:
    try:
        import polars as pl

        df = pl.DataFrame(req.data)
        result = _get_svc().pivot.pivot_data(
            df, req.index, req.columns, req.values, req.aggfunc
        )
        return {"result": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Formula Builder
# ---------------------------------------------------------------------------


@router.post("/excel/formula/build")
async def build_formula(req: BuildFormulaRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        ft = req.formula_type
        p = req.params
        builders = {
            "sum": lambda: svc.formula_builder.sum(p.get("range")),
            "avg": lambda: svc.formula_builder.average(p.get("range")),
            "count": lambda: svc.formula_builder.count(p.get("range")),
            "min": lambda: svc.formula_builder.min(p.get("range")),
            "max": lambda: svc.formula_builder.max(p.get("range")),
            "if": lambda: svc.formula_builder.if_func(
                p.get("condition"), p.get("true_val"), p.get("false_val")
            ),
            "vlookup": lambda: svc.formula_builder.vlookup(
                p.get("lookup_value"),
                p.get("table_range"),
                p.get("col_index"),
                p.get("exact", False),
            ),
            "concat": lambda: svc.formula_builder.concat(p.get("cells", [])),
            "countif": lambda: svc.formula_builder.count_if(
                p.get("range"), p.get("criteria")
            ),
            "sumif": lambda: svc.formula_builder.sum_if(
                p.get("criteria_range"), p.get("criteria"), p.get("sum_range")
            ),
            "index_match": lambda: svc.formula_builder.index_match(
                p.get("lookup_value"),
                p.get("index_range"),
                p.get("match_range"),
                p.get("exact", True),
            ),
        }
        if ft in builders:
            return {"formula": builders[ft]()}
        raise HTTPException(400, detail=f"Unknown formula type: {ft}")
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Document Features
# ---------------------------------------------------------------------------


@router.post("/excel/comments/add")
async def add_comment(req: CommentRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.comments.add(
            wb, req.cell_ref, req.text, req.author, req.sheet_name
        )
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/hyperlinks/add")
async def add_hyperlink(req: HyperlinkRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.hyperlinks.add(wb, req.cell_ref, req.url, req.text, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/images/add")
async def add_image(req: ImageRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.images.add(
            wb, req.image_path, req.cell_ref, req.width, req.height, req.sheet_name
        )
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/named-ranges/add")
async def add_named_range(req: NamedRangeRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.named_ranges.add(wb, req.range_name, req.range_ref, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/named-ranges/list")
async def list_named_ranges(req: FilePathRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        result = _get_svc().named_ranges.list(wb)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/security/protect")
async def protect_workbook(req: SecurityRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.security.protect(
            wb, req.password, req.sheet_name, req.protection_settings
        )
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/security/unprotect")
async def unprotect_workbook(req: SecurityRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.security.unprotect(wb, req.password, req.sheet_name)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/print/setup")
async def print_setup(req: PrintLayoutRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.print_layout.setup(wb, req.sheet_name, req.settings)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/versioning/save")
async def save_version(req: VersionRequest) -> Dict[str, Any]:
    try:
        result = _get_svc().versioning.save_snapshot(
            req.file_path, req.version_path, req.comment
        )
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Sort & Filter
# ---------------------------------------------------------------------------


@router.post("/excel/sort")
async def sort_range(req: SortRangeRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.sort.sort_range(
            wb,
            req.range_address,
            key_column=req.sort_by_column,
            ascending=not req.descending,
            sheet_name=req.sheet_name,
        )
        _save_wb(wb, req.file_path)
        return {"success": result["sorted"], "rows_affected": result["rows"]}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/filter")
async def filter_range(req: FilterRangeRequest) -> Dict[str, Any]:
    try:
        from openpyxl.worksheet.cell_range import CellRange

        wb = _load_wb(req.file_path)
        ws = wb[req.sheet_name]
        cr = CellRange(req.range_address)
        criteria_set = set(req.criteria)
        col_idx = req.field_index - 1
        count = 0
        for row in ws.iter_rows(
            min_row=cr.min_row + 1,
            max_row=cr.max_row,
            min_col=cr.min_col,
            max_col=cr.max_col,
            values_only=True,
        ):
            if len(row) > col_idx and row[col_idx] in criteria_set:
                count += 1
        return {"success": True, "filtered_rows": count}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Versioning
# ---------------------------------------------------------------------------


@router.post("/excel/versioning/snapshot")
async def create_snapshot(req: SnapshotRequest) -> Dict[str, Any]:
    try:
        import os
        import shutil
        from datetime import datetime, timezone

        versions_dir = f"{req.file_path}.versions"
        os.makedirs(versions_dir, exist_ok=True)
        snapshot_path = os.path.join(versions_dir, f"{req.label}.xlsx")
        shutil.copy2(req.file_path, snapshot_path)
        timestamp = datetime.now(timezone.utc).isoformat()
        return {"snapshot_id": req.label, "timestamp": timestamp}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/versioning/diff")
async def diff_snapshots(req: DiffRequest) -> Dict[str, Any]:
    try:
        import os

        import openpyxl
        from openpyxl.utils import get_column_letter

        versions_dir = f"{req.file_path}.versions"
        path_1 = os.path.join(versions_dir, f"{req.snapshot_id_1}.xlsx")
        path_2 = os.path.join(versions_dir, f"{req.snapshot_id_2}.xlsx")
        wb1 = openpyxl.load_workbook(path_1, data_only=True)
        wb2 = openpyxl.load_workbook(path_2, data_only=True)
        differences = []
        all_sheets = set(wb1.sheetnames) | set(wb2.sheetnames)
        for sheet in all_sheets:
            ws1 = wb1[sheet] if sheet in wb1.sheetnames else None
            ws2 = wb2[sheet] if sheet in wb2.sheetnames else None
            max_row = max(ws1.max_row if ws1 else 0, ws2.max_row if ws2 else 0)
            max_col = max(ws1.max_column if ws1 else 0, ws2.max_column if ws2 else 0)
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    val1 = ws1.cell(row=r, column=c).value if ws1 else None
                    val2 = ws2.cell(row=r, column=c).value if ws2 else None
                    if val1 != val2:
                        differences.append(
                            {
                                "cell_address": f"{get_column_letter(c)}{r}",
                                "old_value": val1,
                                "new_value": val2,
                            }
                        )
        return {"differences": differences, "change_count": len(differences)}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# AI Ops
# ---------------------------------------------------------------------------


@router.post("/excel/ai-ops")
async def ai_operations(req: AIOpsRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        wb = _load_wb(req.file_path)
        result = svc.ai_ops.run(wb, req.operation, req.params)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Async Job Operations
# ---------------------------------------------------------------------------


@router.post("/excel/async/convert")
async def async_convert(req: AsyncConvertRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        job_id = svc.submit_async_convert(req.target_format, req.options)
        return {"job_id": job_id, "status": "submitted"}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/async/ocr")
async def async_ocr(req: AsyncOcrRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        job_id = svc.submit_async_ocr(req.image_path, req.options)
        return {"job_id": job_id, "status": "submitted"}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/async/export")
async def async_export(req: AsyncExportRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        job_id = svc.submit_async_export(req.export_format, req.options)
        return {"job_id": job_id, "status": "submitted"}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.get("/excel/async/status/{job_id}")
async def async_job_status(job_id: str) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        return svc.get_job_status(job_id)
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.get("/excel/async/jobs")
async def async_list_jobs(
    status: Optional[str] = Query(None),
) -> List[Dict[str, Any]]:
    try:
        svc = _get_svc()
        return svc.list_jobs(status)
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/async/cancel/{job_id}")
async def async_cancel_job(job_id: str) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        ok = svc.cancel_job(job_id)
        return {"cancelled": ok, "job_id": job_id}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.get("/excel/async/stream/{job_id}")
async def async_job_stream(job_id: str):
    try:
        from fastapi.responses import StreamingResponse

        from common_lib.modules.doc_processing.excel.async_ops import (
            SSEStreamer,
            _get_job_manager,
        )

        job_manager = _get_job_manager()
        stream = SSEStreamer.event_stream(job_id, job_manager)
        return StreamingResponse(stream, media_type="text/event-stream")
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Rendering / Preview
# ---------------------------------------------------------------------------


@router.post("/excel/render/sheet")
async def render_sheet(req: RenderSheetRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        result = _get_svc().renderer.render_sheet_to_html(
            wb,
            req.sheet_name,
            req.max_rows or 100,
            req.max_cols or 20,
            req.include_styles or True,
        )
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/render/markdown")
async def render_markdown(req: RenderMarkdownRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        md = _get_svc().extractor.to_markdown(wb, req.sheet_name, req.max_rows)
        return {"markdown": md}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Security Scanning
# ---------------------------------------------------------------------------


@router.post("/excel/security/scan")
async def security_scan(req: SecurityScanRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        file_result: Dict[str, Any] = {
            "file_type": svc.verify_file_type(req.file_path),
            "malware": svc.scan_for_malware(req.file_path),
            "archive_bomb": svc.check_archive_bomb(req.file_path),
            "zip_slip": svc.check_zip_slip(req.file_path),
        }
        wb_result: Dict[str, Any] = {}
        if req.deep:
            wb = _load_wb(req.file_path)
            wb_result = {
                "dde_links": svc.security.detect_dde_links(wb),
                "data_connections": svc.security.detect_data_connections(wb),
                "ole_objects": svc.security.detect_ole_objects(wb),
                "pii": svc.detect_pii(wb),
            }
        return {"file_scan": file_result, "workbook_scan": wb_result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/security/detect-pii")
async def detect_pii(req: FilePathRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        result = _get_svc().detect_pii(wb)
        return {"pii": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Macro Inspection
# ---------------------------------------------------------------------------


@router.post("/excel/macros/inspect")
async def macros_inspect(req: FilePathRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        result = _get_svc().macro_inventory(wb)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/macros/quarantine")
async def macros_quarantine(req: FilePathRequest) -> Dict[str, Any]:
    try:
        wb = _load_wb(req.file_path)
        result = _get_svc().quarantine_macros(wb)
        _save_wb(wb, req.file_path)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Capability Registries
# ---------------------------------------------------------------------------


@router.get("/excel/capabilities")
async def excel_capabilities(
    registry: Optional[str] = Query(None),
) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        return svc.get_capabilities(registry)
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# OCR Extraction
# ---------------------------------------------------------------------------


@router.post("/excel/ocr/extract")
async def ocr_extract(req: OcrExtractRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        result = svc.ocr.extract_from_image(req.file_path, req.sheet_name or "Sheet1")
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ---------------------------------------------------------------------------
# Collaboration — Locking
# ---------------------------------------------------------------------------


@router.post("/excel/collaboration/lock")
async def collab_lock(req: CollabLockRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        result = svc.lock_manager.lock_cell(
            req.workbook_id, req.cell_ref, req.user_id, req.sheet_name
        )
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.post("/excel/collaboration/unlock")
async def collab_unlock(req: CollabUnlockRequest) -> Dict[str, Any]:
    try:
        svc = _get_svc()
        result = svc.lock_manager.release_lock(req.lock_id, req.user_id)
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))


__all__ = ["router"]
